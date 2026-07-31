"""
Import faculty from xlsx into Supabase advisors table.
Usage: python3 scripts/import-faculty.py
Reads env from apps/web/.env
"""

import os
import sys
import json
import re
import openpyxl
import requests

# Load env from apps/web/.env
env_path = os.path.join(os.path.dirname(__file__), '..', 'apps', 'web', '.env')
if os.path.exists(env_path):
    with open(env_path) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            if '=' not in line:
                continue
            key, _, val = line.partition('=')
            key = key.strip()
            val = val.strip().strip('"').strip("'")
            if key not in os.environ:
                os.environ[key] = val

SUPABASE_URL = os.environ.get('NEXT_PUBLIC_SUPABASE_URL', '')
SUPABASE_KEY = os.environ.get('SUPABASE_SERVICE_ROLE_KEY', '') or os.environ.get('NEXT_PUBLIC_SUPABASE_ANON_KEY', '')
DIRECT_URL = os.environ.get('DIRECT_URL', '')

if not SUPABASE_URL or not SUPABASE_KEY:
    print("Missing Supabase credentials in apps/web/.env")
    sys.exit(1)

# Ensure advisors table exists
if DIRECT_URL:
    try:
        import psycopg2
        conn = psycopg2.connect(DIRECT_URL)
        cur = conn.cursor()
        cur.execute("""
            create table if not exists public.advisors (
              id text primary key,
              name text not null,
              email text unique not null,
              department text default 'CSE',
              assigned_sections text[] default '{}',
              pending_verifications integer default 0,
              phone text,
              office_location text,
              experience integer default 0,
              publications integer default 0,
              created_at timestamptz default now(),
              updated_at timestamptz default now()
            );
        """)

    except Exception as e:
        print(f"Could not create table: {e}")
        print("Ensure advisors table exists in Supabase before running.")
else:
    print("No DIRECT_URL found. Ensure advisors table exists in Supabase.")

XLSX_PATH = '/home/vampire/Downloads/faculty names sheet.xlsx'
if not os.path.exists(XLSX_PATH):
    print(f"File not found: {XLSX_PATH}")
    sys.exit(1)

wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb['Sheet1']

rows = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    sno, name, section_letter = row
    if not name:
        continue
    name = str(name).strip()
    section_letter = str(section_letter).strip().upper() if section_letter else ''
    rows.append((name, section_letter))

print(f"Read {len(rows)} faculty entries")

# Build advisor records: 2 faculty per section → "2A", "2B", ...
advisors = []
for name, section_letter in rows:
    sec = f"2{section_letter}"
    email = re.sub(r'[^a-zA-Z0-9]', '.', name.split('(')[0].strip()).lower()
    email = re.sub(r'\.+', '.', email).strip('.')
    email = f"{email}@citchennai.net"
    
    advisor_id = f"adv-{sec}-{len(advisors) + 1}".lower()

    advisors.append({
        "id": advisor_id,
        "name": name,
        "email": email,
        "department": "CSE",
        "assigned_sections": [sec],
        "pending_verifications": 0,
        "phone": None,
        "office_location": None,
        "experience": 0,
        "publications": 0,
    })

url = f"{SUPABASE_URL}/rest/v1/advisors"
headers = {
    "Content-Type": "application/json",
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
}

# Upsert in batches using POST with merge-duplicates
batch_size = 50
for i in range(0, len(advisors), batch_size):
    batch = advisors[i:i + batch_size]
    resp = requests.post(
        url,
        headers={**headers, "Prefer": "resolution=merge-duplicates"},
        data=json.dumps(batch)
    )
    if resp.status_code in (200, 201):
        print(f"Batch {i // batch_size + 1}: upserted {len(batch)} advisors")
    else:
        print(f"Batch {i // batch_size + 1} failed: {resp.status_code} {resp.text[:200]}")

print(f"Faculty import complete! ({len(advisors)} advisors)")
